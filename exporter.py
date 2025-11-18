"""
Export module for CSV and Google Sheets output.
"""

import csv
import json
import re
from typing import List, Dict, Set, Tuple, Optional
from datetime import datetime
from collections import defaultdict

try:
    import gspread
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False
    gspread = None
    Credentials = None
    Request = None
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Define dummy types for type hints when openpyxl is not available
    Workbook = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    Alignment = None  # type: ignore


def parse_component_labels(component_breakdown: List[str]) -> Tuple[Dict[str, float], Dict[str, float]]:
    """
    Parse component breakdown to extract consigner and investor labels with amounts.
    
    Args:
        component_breakdown: List of component breakdown strings like "Consigner - Label: $amount"
    
    Returns:
        Tuple of (consigners_dict, investors_dict) where keys are labels and values are amounts
    """
    consigners = {}
    investors = {}
    
    for item in component_breakdown:
        # Parse format: "Type - Label: $amount" or "Type: $amount"
        match = re.match(r'^(Consigner|Investor)(?:\s*-\s*([^:]+))?:\s*\$\s*([\d.]+)', item)
        if match:
            comp_type = match.group(1)
            label = match.group(2).strip() if match.group(2) else ""
            amount = float(match.group(3))
            
            if comp_type == "Consigner":
                key = label if label else "Default"
                consigners[key] = consigners.get(key, 0) + amount
            elif comp_type == "Investor":
                key = label if label else "Default"
                investors[key] = investors.get(key, 0) + amount
    
    return consigners, investors


def export_to_csv(breakdowns: List[Dict], output_path: str = "orders_export.csv"):
    """
    Export order breakdowns to CSV file with unique columns for each consigner/investor.
    
    Args:
        breakdowns: List of breakdown dictionaries from rule_engine
        output_path: Path to output CSV file
    """
    if not breakdowns:
        return
    
    # Collect all unique consigner and investor labels
    all_consigner_labels = set()
    all_investor_labels = set()
    all_tax_types = set()  # Collect all unique tax types from Shopify
    breakdowns_with_labels = []
    
    for breakdown in breakdowns:
        component_breakdown = breakdown.get("component_breakdown", [])
        consigners, investors = parse_component_labels(component_breakdown)
        all_consigner_labels.update(consigners.keys())
        all_investor_labels.update(investors.keys())
        
        # Collect tax types from Shopify tax lines
        tax_lines = breakdown.get("tax_lines", []) or []
        for tax_line in tax_lines:
            tax_title = tax_line.get("title", "Tax")
            if tax_title:
                all_tax_types.add(tax_title)
        
        breakdowns_with_labels.append({
            'breakdown': breakdown,
            'consigners': consigners,
            'investors': investors
        })
    
    # Build column headers
    base_columns = [
        "Order ID",
        "Order Number",
        "Date",
        "Customer",
        "Products",
        "Vendor",
        "Product Type",
        "Tags",
        "Collections",
        "Order Total",
        "Total Cost",
        "Revenue",
        "State Taxes",
        "Federal Taxes",
    ]
    
    # Add columns for each unique investor (sorted)
    investor_columns = [f"Investor - {label}" if label != "Default" else "Investor" 
                       for label in sorted(all_investor_labels)]
    
    # Add columns for each unique consigner (sorted)
    consigner_columns = [f"Consigner - {label}" if label != "Default" else "Consigner" 
                        for label in sorted(all_consigner_labels)]
    
    # Add columns for each unique Shopify tax type (sorted)
    shopify_tax_columns = [f"Shopify Tax - {tax_type}" for tax_type in sorted(all_tax_types)]
    
    # Combine all columns
    fieldnames = base_columns + investor_columns + consigner_columns + shopify_tax_columns + [
        "Shopify Tax Breakdown",
        "Component Breakdown",
        "Matched Rules"
    ]
    
    # Write CSV
    with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        # Track totals for numeric columns (exclude new text columns)
        totals = {col: 0.0 for col in fieldnames if col in [
            "Order Total", "Total Cost", "Revenue", "State Taxes", "Federal Taxes"
        ] + investor_columns + consigner_columns + shopify_tax_columns}
        
        # Write data rows
        for item in breakdowns_with_labels:
            breakdown = item['breakdown']
            consigners = item['consigners']
            investors = item['investors']
            
            component_breakdown = breakdown.get("component_breakdown", [])
            breakdown_str = "; ".join(component_breakdown) if component_breakdown else ""
            
            # Build row
            row = {
                "Order ID": breakdown.get("order_id", ""),
                "Order Number": breakdown.get("order_number", ""),
                "Date": breakdown.get("date", ""),
                "Customer": breakdown.get("customer", ""),
                "Products": breakdown.get("products", ""),
                "Vendor": breakdown.get("vendor", ""),
                "Product Type": breakdown.get("product_type", ""),
                "Tags": breakdown.get("tags", ""),
                "Collections": breakdown.get("collections", ""),
                "Order Total": breakdown.get("order_total", 0),
                "Total Cost": breakdown.get("total_cost", 0),
                "Revenue": breakdown.get("revenue", 0),
                "State Taxes": breakdown.get("state_taxes", 0),
                "Federal Taxes": breakdown.get("federal_taxes", 0),
                "Component Breakdown": breakdown_str,
                "Matched Rules": breakdown.get("matched_rules", "")
            }
            
            # Add investor columns
            for label in sorted(all_investor_labels):
                col_name = f"Investor - {label}" if label != "Default" else "Investor"
                amount = investors.get(label, 0)
                row[col_name] = amount
                totals[col_name] = totals.get(col_name, 0) + amount
            
            # Add consigner columns
            for label in sorted(all_consigner_labels):
                col_name = f"Consigner - {label}" if label != "Default" else "Consigner"
                amount = consigners.get(label, 0)
                row[col_name] = amount
                totals[col_name] = totals.get(col_name, 0) + amount
            
            # Add Shopify tax columns
            tax_lines = breakdown.get("tax_lines", []) or []
            shopify_tax_breakdown_str = breakdown.get("shopify_tax_breakdown", [])
            if isinstance(shopify_tax_breakdown_str, list):
                shopify_tax_breakdown_str = "; ".join(shopify_tax_breakdown_str)
            row["Shopify Tax Breakdown"] = shopify_tax_breakdown_str or ""
            
            for tax_type in sorted(all_tax_types):
                col_name = f"Shopify Tax - {tax_type}"
                # Find matching tax line
                tax_amount = 0
                for tax_line in tax_lines:
                    if tax_line.get("title", "") == tax_type:
                        tax_amount = float(tax_line.get("amount", "0"))
                        break
                row[col_name] = tax_amount
                totals[col_name] = totals.get(col_name, 0) + tax_amount
            
            # Update totals for base numeric columns
            for col in ["Order Total", "Total Cost", "Revenue", "State Taxes", "Federal Taxes"]:
                totals[col] += row.get(col, 0) or 0
            
            writer.writerow(row)
        
        # Write totals row
        totals_row = {"Order ID": "TOTAL"}
        for col in fieldnames[1:]:  # Skip Order ID
            if col in totals:
                totals_row[col] = round(totals[col], 2)
            else:
                totals_row[col] = ""
        writer.writerow(totals_row)


def _create_sheet(wb, sheet_name: str, breakdowns: List[Dict], fieldnames: List[str]):
    """Create a sheet with data and totals row."""
    ws = wb.create_sheet(title=sheet_name)
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    numeric_cols = [6, 7, 8, 9, 10, 11, 12]  # Order Total through Consigner columns
    
    for row_idx, breakdown in enumerate(breakdowns, start=2):
        component_breakdown = breakdown.get("component_breakdown", [])
        breakdown_str = "; ".join(component_breakdown) if component_breakdown else ""
        
        values = [
            breakdown.get("order_id", ""),
            breakdown.get("order_number", ""),
            breakdown.get("date", ""),
            breakdown.get("customer", ""),
            breakdown.get("products", ""),
            breakdown.get("order_total", 0),
            breakdown.get("total_cost", 0),
            breakdown.get("revenue", 0),
            breakdown.get("investor", 0),
            breakdown.get("state_taxes", 0),
            breakdown.get("federal_taxes", 0),
            breakdown.get("consigner", 0),
            breakdown_str,
            breakdown.get("matched_rules", "")
        ]
        
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col in numeric_cols:
                cell.number_format = '0.00'
    
    # Add totals row
    totals_row = row_idx + 1
    ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
    
    # Calculate totals for numeric columns
    for col in numeric_cols:
        col_letter = ws.cell(row=1, column=col).column_letter
        formula = f"=SUM({col_letter}2:{col_letter}{row_idx})"
        cell = ws.cell(row=totals_row, column=col, value=formula)
        cell.font = Font(bold=True)
        cell.number_format = '0.00'
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(fieldnames) + 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col).column_letter
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _export_to_csv_fallback(breakdowns: List[Dict], output_path: str):
    """Fallback CSV export if openpyxl is not available."""
    fieldnames = [
        "Order ID",
        "Order Number",
        "Date",
        "Customer",
        "Products",
        "Vendor",
        "Product Type",
        "Tags",
        "Collections",
        "Order Total",
        "Total Cost",
        "Revenue",
        "Investor",
        "State Taxes",
        "Federal Taxes",
        "Consigner",
        "Component Breakdown",
        "Matched Rules"
    ]
    
    with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        totals = {col: 0 for col in fieldnames[9:16]}  # Numeric columns (after text columns)
        
        for breakdown in breakdowns:
            component_breakdown = breakdown.get("component_breakdown", [])
            breakdown_str = "; ".join(component_breakdown) if component_breakdown else ""
            
            row = {
                "Order ID": breakdown.get("order_id", ""),
                "Order Number": breakdown.get("order_number", ""),
                "Date": breakdown.get("date", ""),
                "Customer": breakdown.get("customer", ""),
                "Products": breakdown.get("products", ""),
                "Vendor": breakdown.get("vendor", ""),
                "Product Type": breakdown.get("product_type", ""),
                "Tags": breakdown.get("tags", ""),
                "Collections": breakdown.get("collections", ""),
                "Order Total": breakdown.get("order_total", 0),
                "Total Cost": breakdown.get("total_cost", 0),
                "Revenue": breakdown.get("revenue", 0),
                "Investor": breakdown.get("investor", 0),
                "State Taxes": breakdown.get("state_taxes", 0),
                "Federal Taxes": breakdown.get("federal_taxes", 0),
                "Consigner": breakdown.get("consigner", 0),
                "Component Breakdown": breakdown_str,
                "Matched Rules": breakdown.get("matched_rules", "")
            }
            writer.writerow(row)
            
            # Accumulate totals
            for col in totals.keys():
                totals[col] += row.get(col, 0) or 0
        
        # Write totals row
        totals_row = {"Order ID": "TOTAL"}
        totals_row.update(totals)
        totals_row.update({
            "Order Number": "",
            "Date": "",
            "Customer": "",
            "Products": "",
            "Component Breakdown": "",
            "Matched Rules": ""
        })
        writer.writerow(totals_row)


def export_to_google_sheets(breakdowns: List[Dict], oauth_token_json: str, 
                            spreadsheet_id: Optional[str] = None,
                            client_id: Optional[str] = None,
                            client_secret: Optional[str] = None) -> Dict:
    """
    Export order breakdowns to Google Sheets with monthly tabs.
    
    Args:
        breakdowns: List of breakdown dictionaries from rule_engine
        oauth_token_json: JSON string of OAuth token credentials
        spreadsheet_id: Optional Google Sheets spreadsheet ID (creates new if not provided)
        client_id: OAuth client ID for token refresh
        client_secret: OAuth client secret for token refresh
    
    Returns:
        Dictionary with 'success', 'spreadsheet_id', 'spreadsheet_url', and optional 'error'
    """
    if not GSPREAD_AVAILABLE:
        return {
            'success': False,
            'error': 'gspread library not installed. Please install: pip install gspread google-auth google-auth-oauthlib google-auth-httplib2'
        }
    
    if not breakdowns:
        return {
            'success': False,
            'error': 'No orders to export'
        }
    
    try:
        # Parse OAuth token
        token_data = json.loads(oauth_token_json)
        
        # Create credentials from token
        creds = Credentials.from_authorized_user_info(token_data)
        
        # Track if token was updated (for saving back to DB)
        token_updated = False
        
        # Refresh token if needed
        if creds.expired and creds.refresh_token:
            if client_id and client_secret:
                try:
                    creds.refresh(Request())
                    # Google may add 'openid' scope automatically - this is normal and harmless
                    # Update stored scopes if they've changed (e.g., openid was added)
                    stored_scopes = set(token_data.get('scopes', []))
                    new_scopes = set(creds.scopes) if creds.scopes else set()
                    
                    if new_scopes != stored_scopes:
                        # Scopes changed (likely openid was added) - update stored token
                        token_data['scopes'] = list(creds.scopes) if creds.scopes else []
                        token_data['token'] = creds.token
                        token_updated = True
                except Exception as refresh_error:
                    # If refresh fails due to scope mismatch, try to handle it
                    error_str = str(refresh_error)
                    if 'scope' in error_str.lower() and 'changed' in error_str.lower():
                        # Scope mismatch - user needs to re-authenticate
                        return {
                            'success': False,
                            'error': 'OAuth scopes have changed. Please disconnect and sign in with Google again.'
                        }
                    raise
            else:
                return {
                    'success': False,
                    'error': 'Token expired and refresh failed. Please sign in again.'
                }
        
        # Authenticate with gspread
        gc = gspread.authorize(creds)
        
        # Get or create spreadsheet
        if spreadsheet_id:
            try:
                spreadsheet = gc.open_by_key(spreadsheet_id)
            except gspread.exceptions.SpreadsheetNotFound:
                return {
                    'success': False,
                    'error': f'Spreadsheet not found. Please check the spreadsheet ID or create a new one.'
                }
        else:
            # Create new spreadsheet
            spreadsheet = gc.create(f'Shopify Orders Export - {datetime.now().strftime("%Y-%m-%d")}')
            spreadsheet_id = spreadsheet.id
        
        # Collect all unique consigner and investor labels (same logic as CSV export)
        all_consigner_labels = set()
        all_investor_labels = set()
        all_tax_types = set()  # Collect all unique tax types from Shopify
        breakdowns_with_labels = []
        
        for breakdown in breakdowns:
            component_breakdown = breakdown.get("component_breakdown", [])
            consigners, investors = parse_component_labels(component_breakdown)
            all_consigner_labels.update(consigners.keys())
            all_investor_labels.update(investors.keys())
            
            # Collect tax types from Shopify tax lines
            tax_lines = breakdown.get("tax_lines", []) or []
            for tax_line in tax_lines:
                tax_title = tax_line.get("title", "Tax")
                if tax_title:
                    all_tax_types.add(tax_title)
            
            breakdowns_with_labels.append({
                'breakdown': breakdown,
                'consigners': consigners,
                'investors': investors
            })
        
        # Build column headers (same as CSV)
        base_columns = [
            "Order ID",
            "Order Number",
            "Date",
            "Customer",
            "Products",
            "Vendor",
            "Product Type",
            "Tags",
            "Collections",
            "Order Total",
            "Total Cost",
            "Revenue",
            "State Taxes",
            "Federal Taxes",
        ]
        
        investor_columns = [f"Investor - {label}" if label != "Default" else "Investor" 
                           for label in sorted(all_investor_labels)]
        consigner_columns = [f"Consigner - {label}" if label != "Default" else "Consigner" 
                            for label in sorted(all_consigner_labels)]
        shopify_tax_columns = [f"Shopify Tax - {tax_type}" for tax_type in sorted(all_tax_types)]
        
        fieldnames = base_columns + investor_columns + consigner_columns + shopify_tax_columns + [
            "Shopify Tax Breakdown",
            "Component Breakdown",
            "Matched Rules"
        ]
        
        # Group by month
        monthly_data = organize_by_month(breakdowns)
        
        # Create/update sheets for each month
        for month_key, month_breakdowns in sorted(monthly_data.items()):
            sheet_name = f"{month_key}" if month_key != "Unknown" else "Unknown"
            
            # Try to get existing sheet or create new one
            try:
                worksheet = spreadsheet.worksheet(sheet_name)
                # Clear existing data
                worksheet.clear()
            except gspread.exceptions.WorksheetNotFound:
                worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=len(fieldnames))
            
            # Prepare data rows
            rows = [fieldnames]  # Header row
            
            totals = {col: 0.0 for col in fieldnames if col in [
                "Order Total", "Total Cost", "Revenue", "State Taxes", "Federal Taxes"
            ] + investor_columns + consigner_columns + shopify_tax_columns}
            
            # Filter breakdowns for this month (match by order_id)
            month_order_ids = {b.get('order_id') for b in month_breakdowns}
            month_breakdowns_with_labels = []
            for b in breakdowns_with_labels:
                if b['breakdown'].get('order_id') in month_order_ids:
                    month_breakdowns_with_labels.append(b)
            
            for item in month_breakdowns_with_labels:
                breakdown = item['breakdown']
                consigners = item['consigners']
                investors = item['investors']
                
                component_breakdown = breakdown.get("component_breakdown", [])
                breakdown_str = "; ".join(component_breakdown) if component_breakdown else ""
                
                # Build row
                row = [
                    breakdown.get("order_id", ""),
                    breakdown.get("order_number", ""),
                    breakdown.get("date", ""),
                    breakdown.get("customer", ""),
                    breakdown.get("products", ""),
                    breakdown.get("vendor", ""),
                    breakdown.get("product_type", ""),
                    breakdown.get("tags", ""),
                    breakdown.get("collections", ""),
                    breakdown.get("order_total", 0),
                    breakdown.get("total_cost", 0),
                    breakdown.get("revenue", 0),
                    breakdown.get("state_taxes", 0),
                    breakdown.get("federal_taxes", 0),
                ]
                
                # Add investor columns
                for label in sorted(all_investor_labels):
                    amount = investors.get(label, 0)
                    row.append(amount)
                    totals[f"Investor - {label}" if label != "Default" else "Investor"] += amount
                
                # Add consigner columns
                for label in sorted(all_consigner_labels):
                    amount = consigners.get(label, 0)
                    row.append(amount)
                    totals[f"Consigner - {label}" if label != "Default" else "Consigner"] += amount
                
                # Add Shopify tax columns
                tax_lines = breakdown.get("tax_lines", []) or []
                shopify_tax_breakdown_str = breakdown.get("shopify_tax_breakdown", [])
                if isinstance(shopify_tax_breakdown_str, list):
                    shopify_tax_breakdown_str = "; ".join(shopify_tax_breakdown_str)
                
                for tax_type in sorted(all_tax_types):
                    # Find matching tax line
                    tax_amount = 0
                    for tax_line in tax_lines:
                        if tax_line.get("title", "") == tax_type:
                            tax_amount = float(tax_line.get("amount", "0"))
                            break
                    row.append(tax_amount)
                    totals[f"Shopify Tax - {tax_type}"] += tax_amount
                
                row.append(shopify_tax_breakdown_str or "")
                row.append(breakdown_str)
                row.append(breakdown.get("matched_rules", ""))
                
                # Update totals for base numeric columns
                totals["Order Total"] += breakdown.get("order_total", 0) or 0
                totals["Total Cost"] += breakdown.get("total_cost", 0) or 0
                totals["Revenue"] += breakdown.get("revenue", 0) or 0
                totals["State Taxes"] += breakdown.get("state_taxes", 0) or 0
                totals["Federal Taxes"] += breakdown.get("federal_taxes", 0) or 0
                
                rows.append(row)
            
            # Add totals row
            totals_row = ["TOTAL"]
            for col in fieldnames[1:]:  # Skip Order ID
                if col in totals:
                    totals_row.append(round(totals[col], 2))
                else:
                    totals_row.append("")
            rows.append(totals_row)
            
            # Write to sheet
            worksheet.update('A1', rows, value_input_option='USER_ENTERED')
            
            # Format header row
            worksheet.format('A1:{}1'.format(chr(64 + len(fieldnames))), {
                'backgroundColor': {'red': 0.21, 'green': 0.38, 'blue': 0.57},
                'textFormat': {'foregroundColor': {'red': 1.0, 'green': 1.0, 'blue': 1.0}, 'bold': True},
                'horizontalAlignment': 'CENTER'
            })
            
            # Format totals row
            totals_row_num = len(rows)
            worksheet.format(f'A{totals_row_num}:{chr(64 + len(fieldnames))}{totals_row_num}', {
                'backgroundColor': {'red': 0.85, 'green': 0.88, 'blue': 0.95},
                'textFormat': {'bold': True}
            })
            
            # Note: Google Sheets will auto-resize columns based on content
        
        spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        
        result = {
            'success': True,
            'spreadsheet_id': spreadsheet_id,
            'spreadsheet_url': spreadsheet_url,
            'message': f'Successfully exported {len(breakdowns)} orders to Google Sheets'
        }
        
        # If token was updated (e.g., scopes changed to include openid), return it to save back to DB
        if token_updated:
            result['updated_token'] = token_data
        
        return result
        
    except json.JSONDecodeError:
        return {
            'success': False,
            'error': 'Invalid OAuth token format. Please sign in again.'
        }
    except Exception as e:
        return {
            'success': False,
            'error': f'Export failed: {str(e)}'
        }


def organize_by_month(breakdowns: List[Dict]) -> Dict[str, List[Dict]]:
    """
    Organize breakdowns by month.
    
    Args:
        breakdowns: List of breakdown dictionaries
    
    Returns:
        Dictionary mapping month (YYYY-MM) to list of breakdowns
    """
    monthly_data = defaultdict(list)
    for breakdown in breakdowns:
        date_str = breakdown.get("date", "")
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            month_key = date_obj.strftime("%Y-%m")
            monthly_data[month_key].append(breakdown)
        except (ValueError, TypeError):
            monthly_data["Unknown"].append(breakdown)
    
    return dict(monthly_data)

